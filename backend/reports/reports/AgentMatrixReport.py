import logging
from collections import Counter
from datetime import datetime

from openpyxl import Workbook

from util.xcel_utils import writeUncoloredRow, addFilterAndFreeze, resizeColumnWidth

from reports.ReportBase import ReportBase


class AgentMatrixReport(ReportBase):
    def createWorkbook(self, jobs, controllerData, jobFileName):
        logging.info(f"Creating Agent Matrix Report Workbook")

        workbook = Workbook()
        del workbook["Sheet"]

        allAppAgentVersions = set()
        for host in controllerData.values():
            allAppAgentVersions.update(host["appAgentVersions"])
        allAppAgentVersions = [f"{version[2]}:{version[0]}.{version[1]}" for version in sorted(list(allAppAgentVersions), reverse=True)]

        logging.debug(f"Creating workbook sheet for App Agents")
        appAgentsSheet = workbook.create_sheet(f"Overall - App Agents")
        writeUncoloredRow(appAgentsSheet, 1, ["controller", "application", *allAppAgentVersions])
        # Write Data
        rowIdx = 2
        for host, hostInfo in controllerData.items():
            for application in hostInfo["apm"].values():
                agentVersionMap = Counter(application["appAgentVersions"])
                agentCountRow = []
                for version in allAppAgentVersions:
                    if version in agentVersionMap:
                        agentCountRow.append(agentVersionMap[version])
                    else:
                        agentCountRow.append(0)

                writeUncoloredRow(
                    appAgentsSheet,
                    rowIdx,
                    [
                        hostInfo["controller"].host,
                        application["name"],
                        *agentCountRow,
                    ],
                )
                rowIdx += 1

        addFilterAndFreeze(appAgentsSheet)
        resizeColumnWidth(appAgentsSheet)

        logging.debug(f"Creating workbook sheet for Machine Agents")

        allMachineAgentVersions = set()
        for host in controllerData.values():
            allMachineAgentVersions.update(host["machineAgentVersions"])
        allMachineAgentVersions = [f"{version[0]}.{version[1]}" for version in sorted(list(allMachineAgentVersions), reverse=True)]

        machineAgentsSheet = workbook.create_sheet(f"Overall - Machine Agents")
        writeUncoloredRow(
            machineAgentsSheet,
            1,
            ["controller", "application", *allMachineAgentVersions],
        )
        # Write Data
        rowIdx = 2
        for host, hostInfo in controllerData.items():
            for application in hostInfo["apm"].values():
                agentVersionMap = Counter(application["machineAgentVersions"])
                agentCountRow = []
                for version in allMachineAgentVersions:
                    if version in agentVersionMap:
                        agentCountRow.append(agentVersionMap[version])
                    else:
                        agentCountRow.append(0)

                writeUncoloredRow(
                    machineAgentsSheet,
                    rowIdx,
                    [
                        hostInfo["controller"].host,
                        application["name"],
                        *agentCountRow,
                    ],
                )
                rowIdx += 1

        addFilterAndFreeze(machineAgentsSheet)
        resizeColumnWidth(machineAgentsSheet)

        for agentType in ["appServerAgents", "machineAgents", "dbAgents", "analyticsAgents"]:
            sheet = workbook.create_sheet(f"Individual - {agentType}")

            cols = set()
            for host, hostInfo in controllerData.items():
                for agent in hostInfo[agentType]:
                    cols.update(agent.keys())

            writeUncoloredRow(
                sheet,
                1,
                ["controller", *cols],
            )

            rowIdx = 2
            for host, hostInfo in controllerData.items():
                for agent in hostInfo[agentType]:
                    data = []
                    for field in cols:
                        if field in agent:
                            if "time" in field.lower():
                                try:
                                    data.append(datetime.fromtimestamp(int(agent[field]) / 1000))
                                except ValueError:
                                    data.append(str(agent[field]))
                            else:
                                data.append(str(agent[field]))
                        else:
                            data.append(None)
                    writeUncoloredRow(
                        sheet,
                        rowIdx,
                        [
                            hostInfo["controller"].host,
                            *data,
                        ],
                    )
                    rowIdx += 1

            addFilterAndFreeze(sheet, "B2")
            resizeColumnWidth(sheet)

        logging.debug(f"Saving AgentMatrix Workbook")
        workbook.save(f"output/{jobFileName}/{jobFileName}-AgentMatrix.xlsx")
