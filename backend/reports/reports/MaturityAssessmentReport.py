import logging

from openpyxl import Workbook

from util.xcel_utils import (
    writeUncoloredRow,
    writeColoredRow,
    addFilterAndFreeze,
    resizeColumnWidth,
    writeSummarySheet,
)

from reports.ReportBase import ReportBase


class MaturityAssessmentReport(ReportBase):
    def createWorkbook(self, jobs, controllerData, jobFileName):
        for reportType in ["apm", "brum"]:
            logging.info(f"Creating {reportType} Maturity Assessment Report Workbook")

            # Create Report with Raw Data
            workbook = Workbook()

            summarySheet = workbook["Sheet"]
            summarySheet.title = "Summary"

            analysisSheet = workbook.create_sheet(f"Analysis")

            filteredJobs = [job for job in jobs if job.componentType == reportType]

            jobNameCols = []
            for jobStep in filteredJobs:
                name = type(jobStep).__name__
                jobNameCols.append(name if not name.startswith("OverallAssessment") else "OverallAssessment")
                jobStep.reportData(workbook, controllerData, name)

            # Write Headers
            writeUncoloredRow(
                analysisSheet,
                1,
                [
                    "controller",
                    "componentType",
                    "name",
                    *jobNameCols,
                ],
            )

            rowIdx = 2
            for host, hostInfo in controllerData.items():
                for component in hostInfo[reportType].values():
                    writeColoredRow(
                        analysisSheet,
                        rowIdx,
                        [
                            (hostInfo["controller"].host, None),
                            (reportType, None),
                            (component["name"], None),
                            *[component[jobStep]["computed"] for jobStep in [type(jobStep).__name__ for jobStep in filteredJobs]],
                        ],
                    )
                    rowIdx += 1

            addFilterAndFreeze(analysisSheet)
            resizeColumnWidth(analysisSheet)

            # Now that we have the data , Populate the summary sheet with headers
            writeSummarySheet(summarySheet)

            logging.debug(f"Saving MaturityAssessment-{reportType} Workbook")
            workbook.save(f"output/{jobFileName}/{jobFileName}-MaturityAssessment-{reportType}.xlsx")
