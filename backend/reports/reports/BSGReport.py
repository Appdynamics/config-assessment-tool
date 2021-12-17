import logging

from openpyxl import Workbook

from util.xcel_utils import (
    writeUncoloredRow,
    writeColoredRow,
    addFilterAndFreeze,
    resizeColumnWidth,
    writeSummarySheet,
)

from reports.ReportBase import ReportBase


class BSGReport(ReportBase):
    def createWorkbook(self, jobs, controllerData, jobFileName):
        for reportType in ["apm"]:
            logging.info(f"Creating {reportType} BSG Report Workbook")

            # Create Report with Raw Data
            workbook = Workbook()

            summarySheet = workbook["Sheet"]
            summarySheet.title = "Summary"

            analysisSheet = workbook.create_sheet(f"Analysis")

            filteredJobs = [job for job in jobs if job.componentType == reportType]

            for jobStep in filteredJobs:
                jobStep.reportData(workbook, controllerData, type(jobStep).__name__)

            # Write Headers
            writeUncoloredRow(
                analysisSheet,
                1,
                [
                    "controller",
                    "componentType",
                    "name",
                    *[type(jobStep).__name__ for jobStep in filteredJobs],
                ],
            )

            rowIdx = 2
            for host, hostInfo in controllerData.items():
                for component in hostInfo[reportType].values():
                    writeColoredRow(
                        analysisSheet,
                        rowIdx,
                        [
                            (hostInfo["controller"].host, None),
                            ("apm", None),
                            (component["name"], None),
                            *[component[jobStep]["computed"] for jobStep in [type(jobStep).__name__ for jobStep in filteredJobs]],
                        ],
                    )
                    rowIdx += 1

            addFilterAndFreeze(analysisSheet)
            resizeColumnWidth(analysisSheet)

            # Now that we have the data , Populate the summary sheet with headers
            writeSummarySheet(summarySheet)

            logging.debug(f"Saving BSG Report Workbook")
            workbook.save(f"output/{jobFileName}/{jobFileName}-BSGReport-{reportType}.xlsx")
