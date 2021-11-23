import logging

from openpyxl import Workbook

from util.xcel_utils import (
    writeUncoloredRow,
    writeColoredRow,
    addFilterAndFreeze,
    resizeColumnWidth,
    writeSummarySheet,
)

from reports.ReportBase import ReportBase


class ApmReport(ReportBase):
    def createWorkbook(self, jobs, controllerData, jobFileName):
        logging.info(f"Creating Analysis Report Workbook")

        # Create Report with Raw Data
        workbook = Workbook()

        summarySheet = workbook["Sheet"]
        summarySheet.title = "Summary"

        analysisSheet = workbook.create_sheet(f"Analysis")

        for jobStep in jobs:
            jobStep.reportData(workbook, controllerData, type(jobStep).__name__)

        # Write Headers
        writeUncoloredRow(
            analysisSheet,
            1,
            [
                "controller",
                "componentType",
                "application",
                *[type(jobStep).__name__ for jobStep in jobs],
            ],
        )

        rowIdx = 2
        for host, hostInfo in controllerData.items():
            for component in hostInfo["apm"].values():
                writeColoredRow(
                    analysisSheet,
                    rowIdx,
                    [
                        (hostInfo["controller"].host, None),
                        ("apm", None),
                        (component["name"], None),
                        *[component[jobStep]["computed"] for jobStep in [type(jobStep).__name__ for jobStep in jobs]],
                    ],
                )
                rowIdx += 1

        addFilterAndFreeze(analysisSheet)
        resizeColumnWidth(analysisSheet)

        # Now that we have the data , Populate the summary sheet with headers
        writeSummarySheet(summarySheet)

        logging.debug(f"Saving Analysis Report Workbook")
        workbook.save(f"output/{jobFileName}/{jobFileName}-ApmReport.xlsx")
