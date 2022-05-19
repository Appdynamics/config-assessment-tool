import logging
from datetime import datetime

from openpyxl import Workbook
from output.ReportBase import ReportBase
from util.xcel_utils import Color, addFilterAndFreeze, resizeColumnWidth, writeRow, writeUncoloredRow


class LicenseReport(ReportBase):
    def createWorkbook(self, jobs, controllerData, jobFileName):
        logging.info(f"Creating License Report Workbook")

        workbook = Workbook()
        del workbook["Sheet"]

        logging.debug(f"Creating workbook sheet for App Agents")
        licenseSheet = workbook.create_sheet(f"License")
        writeUncoloredRow(licenseSheet, 1, ["controller", "licenseType", "isLicensed", "peakUsage", "numOfProvisionedLicense", "expirationDate"])

        # Write Data
        rowIdx = 2
        for host, hostInfo in controllerData.items():
            for licenseType, licenseData in hostInfo["licenseUsage"].items():
                color = None
                if licenseData is not None and licenseData["isLicensed"] and float(licenseData["numOfProvisionedLicense"]) != 0:
                    if float(licenseData["peakUsage"]) / float(licenseData["numOfProvisionedLicense"]) > 0.75:
                        color = Color.yellow
                    if float(licenseData["peakUsage"]) / float(licenseData["numOfProvisionedLicense"]) > 0.95:
                        color = Color.red

                writeRow(
                    licenseSheet,
                    rowIdx,
                    [
                        (hostInfo["controller"].host, color),
                        (licenseType.split("LicenseProperties")[0], color),
                        (bool(licenseData["isLicensed"]) if licenseData is not None else False, color),
                        (licenseData["peakUsage"] if licenseData is not None else None, color),
                        (licenseData["numOfProvisionedLicense"] if licenseData is not None else None, color),
                        (
                            datetime.fromtimestamp(int(licenseData["expirationDate"]) / 1000)
                            if licenseData is not None and licenseData["expirationDate"] is not None
                            else None,
                            color,
                        ),
                    ],
                )
                rowIdx += 1

        addFilterAndFreeze(licenseSheet)
        resizeColumnWidth(licenseSheet)

        logging.debug(f"Saving License Workbook")
        workbook.save(f"output/{jobFileName}/{jobFileName}-License.xlsx")
