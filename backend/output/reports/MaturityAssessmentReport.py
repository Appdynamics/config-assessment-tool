import logging

from openpyxl import Workbook
from output.ReportBase import ReportBase
from util.excel_utils import addFilterAndFreeze, resizeColumnWidth, writeColoredRow, writeSummarySheet, writeUncoloredRow


class MaturityAssessmentReport(ReportBase):
    def createWorkbook(self, jobs, controllerData, jobFileName):
        for reportType in ["apm", "brum", "mrum"]:
            logging.info(f"Creating {reportType} Maturity Assessment Report Workbook")

            # Create Report with Raw Data
            workbook = Workbook()

            summarySheet = workbook["Sheet"]
            summarySheet.title = "Summary"

            analysisSheet = workbook.create_sheet(f"Analysis")

            filteredJobs = [job for job in jobs if job.componentType == reportType]

            jobNameCols = []
            for jobStep in filteredJobs:
                name = type(jobStep).__name__
                jobNameCols.append(name if not name.startswith("OverallAssessment") else "OverallAssessment")
                jobStep.reportData(workbook, controllerData, name)

            data_header = [
                "controller",
                "componentType",
                "name",
                "applicationId",
                *jobNameCols,
            ]

            if reportType == "apm": # add desc header after name
                data_header.insert(4, "description")

            # Write Headers
            writeUncoloredRow(
                analysisSheet,
                1,
                data_header
            )

            rowIdx = 2
            for host, hostInfo in controllerData.items():
                for component in hostInfo[reportType].values():

                    data_row = [
                        (hostInfo["controller"].host, None),
                        (reportType, None),
                        (component["name"], None),
                        (component["applicationId"] if reportType == "mrum" else component["id"], None),
                        *[component[jobStep]["computed"] for jobStep in [type(jobStep).__name__ for jobStep in filteredJobs]],
                    ]

                    if reportType == "apm": # add desc after name
                        data_row.insert(4, (component["description"], None))

                    writeColoredRow(
                        analysisSheet,
                        rowIdx,
                        data_row
                    )
                    rowIdx += 1

            addFilterAndFreeze(analysisSheet, "E2")
            resizeColumnWidth(analysisSheet)

            # Now that we have the data , Populate the summary sheet with headers
            writeSummarySheet(summarySheet)

            logging.debug(f"Saving MaturityAssessment-{reportType} Workbook")
            workbook.save(f"output/{jobFileName}/{jobFileName}-MaturityAssessment-{reportType}.xlsx")
