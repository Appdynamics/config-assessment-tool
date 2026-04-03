import logging
import os

from openpyxl import Workbook
from backend.output.ReportBase import ReportBase
from backend.util.excel_utils import addFilterAndFreeze, resizeColumnWidth


class RawMaturityAssessmentReport(ReportBase):
    def createWorkbook(self, jobs, controllerData, jobFileName, output_dir="output"):
        for reportType in ["apm", "brum", "mrum"]:
            logging.info(f"Creating {reportType} Raw Maturity Assessment Report Workbook")

            # Create Report with Raw Maturity Assessment Report
            workbook = Workbook()

            filteredJobs = [job for job in jobs if job.componentType == reportType]

            if filteredJobs:
                del workbook["Sheet"]
            else:
                summarySheet = workbook["Sheet"]
                summarySheet.title = "Summary"

            for jobStep in filteredJobs:
                jobStep.reportData(workbook, controllerData, type(jobStep).__name__, False, False)

            analysisSheet = workbook.create_sheet("Analysis")
            addFilterAndFreeze(analysisSheet, "E2")
            resizeColumnWidth(analysisSheet)

            logging.debug(f"Saving MaturityAssessmentRaw-{reportType} Workbook")
            save_path = os.path.join(output_dir, jobFileName, f"{jobFileName}-MaturityAssessmentRaw-{reportType}.xlsx")
            workbook.save(save_path)
