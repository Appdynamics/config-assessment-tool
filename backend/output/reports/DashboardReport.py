import logging
import os

from datetime import datetime
from openpyxl import Workbook

from backend.output.ReportBase import ReportBase
from backend.util.excel_utils import Color, addFilterAndFreeze, resizeColumnWidth, writeRow, writeUncoloredRow


class DashboardReport(ReportBase):
    def createWorkbook(self, jobs, controllerData, jobFileName, output_dir="output"):
        logging.info("Creating Dashboard Report Workbook")

        workbook = Workbook()
        del workbook["Sheet"]

        logging.debug(f"Creating workbook sheet for Dashboards")
        dashboardSheet = workbook.create_sheet(f"Dashboards")
        writeUncoloredRow(dashboardSheet, 1, ["controller", "dashboardName"])

        # Write Data
        rowIdx = 2
        for host, hostInfo in controllerData.items():
            for dashboard in hostInfo["exportedDashboards"]:
                color = None

                writeRow(
                    dashboardSheet,
                    rowIdx,
                    [
                        (hostInfo["controller"].host, color),
                        (dashboard["name"], color),
                    ],
                )
                rowIdx += 1

        addFilterAndFreeze(dashboardSheet, "C2")
        resizeColumnWidth(dashboardSheet)

        logging.debug(f"Saving Dashboard Workbook")
        save_path = os.path.join(output_dir, jobFileName, f"{jobFileName}-Dashboards.xlsx")
        workbook.save(save_path)
