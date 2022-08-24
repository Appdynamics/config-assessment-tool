import json
import logging
from collections import Counter
from datetime import datetime

from openpyxl import Workbook
from output.ReportBase import ReportBase
from util.excel_utils import addFilterAndFreeze, resizeColumnWidth, writeUncoloredRow


class AgentMatrixReport(ReportBase):
    def createWorkbook(self, jobs, controllerData, jobFileName):
        logging.info(f"Creating Agent Matrix Report Workbook")

        workbook = Workbook()
        del workbook["Sheet"]

        allAppAgentVersions = set()
        for host in controllerData.values():
            allAppAgentVersions.update(host["appAgentVersions"])
        allAppAgentVersions = [f"{version[2]}:{version[0]}.{version[1]}" for version in sorted(list(allAppAgentVersions), reverse=True)]

        logging.debug(f"Creating workbook sheet for App Agents")
        appAgentsSheet = workbook.create_sheet(f"Overall - App Agents")
        writeUncoloredRow(appAgentsSheet, 1, ["controller", "application", *allAppAgentVersions])
        # Write Data
        rowIdx = 2
        for host, hostInfo in controllerData.items():
            for application in hostInfo["apm"].values():
                agentVersionMap = Counter(application["appAgentVersions"])
                agentCountRow = []
                for version in allAppAgentVersions:
                    if version in agentVersionMap:
                        agentCountRow.append(agentVersionMap[version])
                    else:
                        agentCountRow.append(0)

                writeUncoloredRow(
                    appAgentsSheet,
                    rowIdx,
                    [
                        hostInfo["controller"].host,
                        application["name"],
                        *agentCountRow,
                    ],
                )
                rowIdx += 1

        addFilterAndFreeze(appAgentsSheet)
        resizeColumnWidth(appAgentsSheet)

        logging.debug(f"Creating workbook sheet for Machine Agents")

        allMachineAgentVersions = set()
        for host in controllerData.values():
            allMachineAgentVersions.update(host["machineAgentVersions"])
        allMachineAgentVersions = [f"{version[0]}.{version[1]}" for version in sorted(list(allMachineAgentVersions), reverse=True)]

        machineAgentsSheet = workbook.create_sheet(f"Overall - Machine Agents")
        writeUncoloredRow(
            machineAgentsSheet,
            1,
            ["controller", "application", *allMachineAgentVersions],
        )
        # Write Data
        rowIdx = 2
        for host, hostInfo in controllerData.items():
            for application in hostInfo["apm"].values():
                agentVersionMap = Counter(application["machineAgentVersions"])
                agentCountRow = []
                for version in allMachineAgentVersions:
                    if version in agentVersionMap:
                        agentCountRow.append(agentVersionMap[version])
                    else:
                        agentCountRow.append(0)

                writeUncoloredRow(
                    machineAgentsSheet,
                    rowIdx,
                    [
                        hostInfo["controller"].host,
                        application["name"],
                        *agentCountRow,
                    ],
                )
                rowIdx += 1

        addFilterAndFreeze(machineAgentsSheet)
        resizeColumnWidth(machineAgentsSheet)

        for agentType in ["appServerAgents", "machineAgents", "dbAgents", "analyticsAgents"]:
            sheet = workbook.create_sheet(f"Individual - {agentType}")

            cols = []
            for host, hostInfo in controllerData.items():
                for agent in hostInfo[agentType]:
                    for key in agent.keys():
                        if key not in cols:
                            cols.append(key)

            machineAgentCols = cols.copy()
            appServerAgentCols = cols.copy()

            machineAgentCols.extend(
                [
                    "simEnabled",
                    "historical",
                    "reportingData",
                    "Physical Cores",
                    "vCPUs",
                    "OS|Architecture",  # in properties section
                    "Bios|Version",
                    "AppDynamics|Agent|Install Directory",
                    "OS|Kernel|Release",
                    "AppDynamics|Agent|Build Number",
                    "AppDynamics|Machine Type",
                    "OS|Kernel|Name",
                    "AppDynamics|Agent|Machine Info",
                    "Total|CPU|Logical Processor Count",
                    "AppDynamics|Agent|JVM Info",
                    "tags",
                ]
            )

            appServerAgentCols.extend(
                [
                    "reportingData",
                    "installDir",
                    "agentVersion",
                    "latestAgentRuntime",
                    "metadata",
                ]
            )

            rowIdx = 2
            for host, hostInfo in controllerData.items():
                for agent in hostInfo[agentType]:
                    data = []
                    for field in cols:
                        if field in agent:
                            if "time" in field.lower():
                                try:
                                    data.append(datetime.fromtimestamp(int(agent[field]) / 1000))
                                except ValueError:
                                    data.append(str(agent[field]))
                            else:
                                data.append(str(agent[field]))
                        else:
                            data.append(None)

                    if agentType == "machineAgents":
                        if agent["hostName"] in hostInfo["servers"]:
                            server = hostInfo["servers"][agent["hostName"]]

                            # use id shown in UI
                            data[cols.index("machineId")] = server["id"]

                            data.append(server["simEnabled"] if "simEnabled" in server else "")
                            data.append(server["historical"] if "historical" in server else "")
                            data.append(server["availability"] if "availability" in server else "")
                            data.append(server["physicalCores"] if "physicalCores" in server else "")
                            data.append(server["virtualCores"] if "virtualCores" in server else "")

                            if "properties" in server:
                                for additionalMachineAgentField in machineAgentCols[machineAgentCols.index("OS|Architecture") :]:
                                    data.append(
                                        server["properties"][additionalMachineAgentField]
                                        if additionalMachineAgentField in server["properties"]
                                        else ""
                                    )

                            data.append(json.dumps(server["tags"]) if "tags" in server else "")
                        else:
                            # simEnabled
                            data.append(False)
                            # historical
                            data.append(None)
                            # reportingData
                            if agent["applicationIds"]:
                                availability = hostInfo["nodeMachineIdMachineAgentAvailabilityMap"].get(agent["machineId"], None)
                                data.append(availability)
                            # Physical Cores
                            data.append(None)
                            # vCPUs
                            data.append(None)

                    elif agentType == "appServerAgents":
                        # reportingData
                        availability = hostInfo["nodeIdAppAgentAvailabilityMap"].get(agent["applicationComponentNodeId"], None)
                        data.append(availability)
                        metadata = hostInfo["nodeIdMetaInfoMap"].get(agent["applicationComponentNodeId"], None)
                        if metadata:
                            metaInfo = [
                                {
                                    "name": info["name"],
                                    "value": info["value"],
                                }
                                for info in metadata["applicationComponentNode"]["metaInfo"]
                            ]
                            data.append(metadata["applicationComponentNode"]["appAgent"]["installDir"])
                            data.append(metadata["applicationComponentNode"]["appAgent"]["agentVersion"])
                            data.append(metadata["applicationComponentNode"]["appAgent"]["latestAgentRuntime"])
                            data.append(json.dumps(metaInfo))

                    writeUncoloredRow(
                        sheet,
                        rowIdx,
                        [
                            hostInfo["controller"].host,
                            *data,
                        ],
                    )
                    rowIdx += 1

            if agentType == "machineAgents":
                writeUncoloredRow(
                    sheet,
                    1,
                    ["controller", *machineAgentCols],
                )
            elif agentType == "appServerAgents":
                writeUncoloredRow(
                    sheet,
                    1,
                    ["controller", *appServerAgentCols],
                )
            else:
                writeUncoloredRow(
                    sheet,
                    1,
                    ["controller", *cols],
                )

            addFilterAndFreeze(sheet, "B2")
            resizeColumnWidth(sheet)

        logging.debug(f"Saving AgentMatrix Workbook")
        workbook.save(f"output/{jobFileName}/{jobFileName}-AgentMatrix.xlsx")
