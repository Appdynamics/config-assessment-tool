import logging
import string
from enum import Enum
from typing import Any, List

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import PatternFill
from openpyxl.utils.exceptions import IllegalCharacterError
from openpyxl.worksheet.worksheet import Worksheet


class Color(Enum):
    """Colors used to denote maturity level in xlsx sheet."""

    platinum = PatternFill(start_color="FFA890F7", end_color="FFA890F7", fill_type="solid")
    gold = PatternFill(start_color="FFFFD700", end_color="FFFFD700", fill_type="solid")
    silver = PatternFill(start_color="FFC0C0C0", end_color="FFC0C0C0", fill_type="solid")
    bronze = PatternFill(start_color="FFcd7f32", end_color="FFcd7f32", fill_type="solid")
    """Colors used to denote issues in other reports."""
    green = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")
    red = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    yellow = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
    white = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")


def writeRow(sheet: Worksheet, rowIdx: int, data: [Any]):
    if all(value for value in data if isinstance(value, tuple) and isinstance(value[1], Color)):
        writeColoredRow(sheet, rowIdx, data)
    else:
        writeUncoloredRow(sheet, rowIdx, data)


def writeColoredRow(sheet: Worksheet, rowIdx: int, data: [(Any, Color)]):
    """Write row of data at given rowIdx starting from colIdx A."""
    for (data, color), colIdx in zip(data, string.ascii_uppercase[: len(data)]):
        sheet[f"{colIdx}{rowIdx}"] = data
        if color is not None:
            sheet[f"{colIdx}{rowIdx}"].fill = color.value


def writeUncoloredRow(sheet: Worksheet, rowIdx: int, data: [Any]):
    """Write row of data at given rowIdx starting from colIdx A. Typically used for writing headers."""
    for cell, colIdx in zip(data, string.ascii_uppercase[: len(data)]):
        try:
            sheet[f"{colIdx}{rowIdx}"] = cell
        except IllegalCharacterError:
            logging.warning(f"illegal character detected in cell, will scrub {cell}")
            cell = ILLEGAL_CHARACTERS_RE.sub(r'', cell)
            logging.warning(f"scrubbed cell: {cell}")
            sheet[f"{colIdx}{rowIdx}"] = cell


def createSheet(workbook: Workbook, sheetName: str, headers: List[Any], rows: List[List[Any]]):
    sheet = workbook.create_sheet(sheetName)
    writeRow(sheet, 1, headers)
    for idx, row in enumerate(rows):
        writeRow(sheet, idx + 1, row)


def writeSummarySheet(summarySheet: Worksheet):
    """Summarize the data in the report on the summary sheet."""
    writeColoredRow(
        summarySheet,
        1,
        [
            ("", None),
            ("bronze", Color.bronze),
            ("silver", Color.silver),
            ("gold", Color.gold),
            ("platinum", Color.platinum),
        ],
    )
    writeUncoloredRow(
        summarySheet,
        2,
        [
            "# of Apps",
            '=COUNTIF(INDEX(Analysis!$1:$1048576,0,MATCH("OverallAssessment",Analysis!$1:$1,0)), B1)',
            '=COUNTIF(INDEX(Analysis!$1:$1048576,0,MATCH("OverallAssessment",Analysis!$1:$1,0)), C1)',
            '=COUNTIF(INDEX(Analysis!$1:$1048576,0,MATCH("OverallAssessment",Analysis!$1:$1,0)), D1)',
            '=COUNTIF(INDEX(Analysis!$1:$1048576,0,MATCH("OverallAssessment",Analysis!$1:$1,0)), E1)',
        ],
    )
    writeUncoloredRow(
        summarySheet,
        3,
        [
            "% of Apps",
            "=ROUND(B2/(COUNTA(Analysis!$C:$C)-1)*100, 1)",
            "=ROUND(C2/(COUNTA(Analysis!$C:$C)-1)*100, 1)",
            "=ROUND(D2/(COUNTA(Analysis!$C:$C)-1)*100, 1)",
            "=ROUND(E2/(COUNTA(Analysis!$C:$C)-1)*100, 1)",
        ],
    )


def resizeColumnWidth(sheet: Worksheet):
    """Resize columns to max width of cell per column."""
    headerFilterArrowPadding = 5
    dims = {}
    for row in sheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        sheet.column_dimensions[col].width = value + headerFilterArrowPadding


def addFilterAndFreeze(sheet: Worksheet, freezePane: str = "C2"):
    """Add filter on headers and freeze the first 2 columns and 1 row."""
    sheet.auto_filter.ref = sheet.dimensions
    # Freeze controller and application columns
    sheet.freeze_panes = freezePane
