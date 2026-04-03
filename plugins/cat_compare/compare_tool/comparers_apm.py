# compare_tool/comparers.py

import logging
from typing import Dict, Tuple, Any

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from .excel_io import get_key_column

logger = logging.getLogger(__name__)

# Color fills (copied from original script)
red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
added_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')

"""
comparers_apm.py
----------------
This module contains the logic for comparing APM (Application Performance Management) data.

Purpose:
- Compares data between the "previous" and "current" APM Excel files.
- Handles sheet-specific comparisons using defined comparer functions.
- Logs warnings for missing sheets or undefined comparers.

Key Features:
- Iterates through all sheets in the "current" workbook.
- Uses `SHEET_COMPARERS_APM` to apply sheet-specific comparison logic.
- Saves the comparison results to the specified output file.
- Logs detailed information about the comparison process.

Key Functions:
- `compare_files_other_sheets_apm`: Compares all sheets except "Summary".
"""

def compare_analysis(ws_previous, ws_current):
    try:
        # Define the columns for comparison
        columns = {
            'AppAgentsAPM': None,
            'MachineAgentsAPM': None,
            'BusinessTransactionsAPM': None,
            'BackendsAPM': None,
            'OverheadAPM': None,
            'ServiceEndpointsAPM': None,
            'ErrorConfigurationAPM': None,
            'HealthRulesAndAlertingAPM': None,
            'DataCollectorsAPM': None,
            'DashboardsAPM': None,
            'OverallAssessment': None
        }

        # Retrieve column indices
        for column in columns.keys():
            col_idx_prev = get_key_column(ws_previous, column)
            col_idx_curr = get_key_column(ws_current, column)
            if col_idx_prev is None or col_idx_curr is None:
                logging.error(f"The '{column}' column is missing in one of the sheets. Cannot proceed with comparison.")
                return
            columns[column] = (col_idx_prev, col_idx_curr)

        # Retrieve key column indices
        ctrl_col_prev = get_key_column(ws_previous, 'controller')
        ctrl_col_curr = get_key_column(ws_current, 'controller')
        name_col_prev = get_key_column(ws_previous, 'name')
        name_col_curr = get_key_column(ws_current, 'name')

        if name_col_prev is None or name_col_curr is None:
            logging.error("The 'name' column is missing in one of the sheets. Cannot proceed with comparison.")
            return

        if ctrl_col_prev is None or ctrl_col_curr is None:
            logging.error("The 'controller' column is missing in one of the sheets. This might affect the comparison.")
            return

        previous_data = {}
        current_data = {}

        # Read previous data
        for row in ws_previous.iter_rows(min_row=2, values_only=False):
            name_value = row[name_col_prev - 1].value
            ctrl_value = row[ctrl_col_prev - 1].value
            key = (name_value, ctrl_value)
            if name_value and ctrl_value:
                previous_data[key] = row

        # Read current data
        for row in ws_current.iter_rows(min_row=2, values_only=False):
            name_value = row[name_col_curr - 1].value
            ctrl_value = row[ctrl_col_curr - 1].value
            key = (name_value, ctrl_value)
            if name_value and ctrl_value:
                current_data[key] = row

        # Compare previous data with current data
        for key, previous_row in previous_data.items():
            current_row = current_data.get(key)
            if current_row:
                for column, (col_idx_prev, col_idx_curr) in columns.items():
                    previous_value = previous_row[col_idx_prev - 1].value
                    current_value = current_row[col_idx_curr - 1].value
                    cell_output = ws_current.cell(row=current_row[0].row, column=col_idx_curr)

                    if previous_value == current_value:
                        continue

                    try:
                        # Comparison logic based on ranking
                        ranking = {'bronze': 1, 'silver': 2, 'gold': 3, 'platinum': 4}
                        prev_rank = ranking.get(str(previous_value).lower(), 0)
                        curr_rank = ranking.get(str(current_value).lower(), 0)

                        if curr_rank > prev_rank:
                            cell_output.fill = green_fill
                            cell_output.value = f"{previous_value} → {current_value} (Upgraded)"
                        elif curr_rank < prev_rank:
                            cell_output.fill = red_fill
                            cell_output.value = f"{previous_value} → {current_value} (Downgraded)"
                    except ValueError:
                        logging.error(
                            f"Invalid ranking value encountered for column '{column}': "
                            f"Previous={previous_value}, Current={current_value}"
                        )

        # Add new entries in the current sheet
        for key, current_row in current_data.items():
            if key not in previous_data:
                row_index = ws_current.max_row + 1
                for col_num, cell in enumerate(current_row, 1):
                    new_cell = ws_current.cell(row=row_index, column=col_num, value=cell.value)
                    new_cell.fill = added_fill

    except Exception as e:
        logging.error(f"Error in compare_analysis: {e}", exc_info=True)
        raise


def compare_appagentsapm(ws_previous, ws_current):
    try:
        columns = {
            'metricLimitNotHit': None,
            'percentAgentsLessThan1YearOld': None,
            'percentAgentsLessThan2YearsOld': None,
            'percentAgentsReportingData': None,
            'percentAgentsRunningSameVersion': None
        }

        for column in columns.keys():
            col_idx_prev = get_key_column(ws_previous, column)
            col_idx_curr = get_key_column(ws_current, column)
            if col_idx_prev is None or col_idx_curr is None:
                logging.error(f"The '{column}' column is missing in one of the sheets. Cannot proceed with comparison.")
                return
            columns[column] = (col_idx_prev, col_idx_curr)

        app_col_prev = get_key_column(ws_previous, 'application')
        app_col_curr = get_key_column(ws_current, 'application')
        ctrl_col_prev = get_key_column(ws_previous, 'controller')
        ctrl_col_curr = get_key_column(ws_current, 'controller')

        if app_col_prev is None or app_col_curr is None:
            logging.error("The 'application' column is missing in one of the sheets. Cannot proceed with comparison.")
            return

        if ctrl_col_prev is None or ctrl_col_curr is None:
            logging.error("The 'controller' column is missing in one of the sheets. This might affect the comparison.")
            return

        previous_data = {}
        current_data = {}

        for row in ws_previous.iter_rows(min_row=2, values_only=False):
            app_value = row[app_col_prev - 1].value
            ctrl_value = row[ctrl_col_prev - 1].value
            key = (app_value, ctrl_value)
            if app_value and ctrl_value:
                previous_data[key] = row

        for row in ws_current.iter_rows(min_row=2, values_only=False):
            app_value = row[app_col_curr - 1].value
            ctrl_value = row[ctrl_col_curr - 1].value
            key = (app_value, ctrl_value)
            if app_value and ctrl_value:
                current_data[key] = row

        for key, previous_row in previous_data.items():
            current_row = current_data.get(key)
            if current_row:
                for column, (col_idx_prev, col_idx_curr) in columns.items():
                    previous_value = previous_row[col_idx_prev - 1].value
                    current_value = current_row[col_idx_curr - 1].value
                    cell_output = ws_current.cell(row=current_row[0].row, column=col_idx_curr)

                    if previous_value == current_value:
                        continue

                    if column == 'metricLimitNotHit':
                        prev_value_str = str(previous_value).strip().upper()
                        curr_value_str = str(current_value).strip().upper()

                        logging.info(f"Comparing {column}: Previous={prev_value_str}, Current={curr_value_str}")

                        if prev_value_str == "FALSE" and curr_value_str == "TRUE":
                            cell_output.fill = green_fill
                            cell_output.value = f"{previous_value} → {current_value} (Improved)"
                        elif prev_value_str == "TRUE" and curr_value_str == "FALSE":
                            cell_output.fill = red_fill
                            cell_output.value = f"{previous_value} → {current_value} (Declined)"
                        else:
                            cell_output.fill = red_fill
                            cell_output.value = f"{previous_value} → {current_value} (Changed)"
                    else:
                        try:
                            prev_value_num = float(previous_value)
                            curr_value_num = float(current_value)
                            formatted_prev_value = f"{prev_value_num:.2f}"
                            formatted_curr_value = f"{curr_value_num:.2f}"

                            if curr_value_num > prev_value_num:
                                cell_output.fill = green_fill
                                cell_output.value = f"{formatted_prev_value} → {formatted_curr_value} (Improved)"
                            else:
                                cell_output.fill = red_fill
                                cell_output.value = f"{formatted_prev_value} → {formatted_curr_value} (Declined)"
                        except ValueError:
                            logging.error(
                                f"Non-numeric value encountered for column '{column}': "
                                f"Previous={previous_value}, Current={current_value}"
                            )

        for key, current_row in current_data.items():
            if key not in previous_data:
                row_index = ws_current.max_row + 1
                for col_num, cell in enumerate(current_row, 1):
                    new_cell = ws_current.cell(row=row_index, column=col_num, value=cell.value)
                    new_cell.fill = added_fill

    except Exception as e:
        logging.error(f"Error in compare_appagentsapm: {e}", exc_info=True)
        raise


def compare_machineagentsapm(ws_previous, ws_current):
    try:
        columns = {
            'percentAgentsLessThan1YearOld': None,
            'percentAgentsLessThan2YearsOld': None,
            'percentAgentsReportingData': None,
            'percentAgentsRunningSameVersion': None,
            'percentAgentsInstalledAlongsideAppAgents': None
        }

        for column in columns.keys():
            col_idx_prev = get_key_column(ws_previous, column)
            col_idx_curr = get_key_column(ws_current, column)
            if col_idx_prev is None or col_idx_curr is None:
                logging.error(f"The '{column}' column is missing in one of the sheets. Cannot proceed with comparison.")
                return
            columns[column] = (col_idx_prev, col_idx_curr)

        app_col_prev = get_key_column(ws_previous, 'application')
        app_col_curr = get_key_column(ws_current, 'application')
        ctrl_col_prev = get_key_column(ws_previous, 'controller')
        ctrl_col_curr = get_key_column(ws_current, 'controller')

        if app_col_prev is None or app_col_curr is None:
            logging.error("The 'application' column is missing in one of the sheets. Cannot proceed with comparison.")
            return

        if ctrl_col_prev is None or ctrl_col_curr is None:
            logging.error("The 'controller' column is missing in one of the sheets. This might affect the comparison.")
            return

        previous_data = {}
        current_data = {}

        for row in ws_previous.iter_rows(min_row=2, values_only=False):
            app_value = row[app_col_prev - 1].value
            ctrl_value = row[ctrl_col_prev - 1].value
            key = (app_value, ctrl_value)
            if app_value and ctrl_value:
                previous_data[key] = row

        for row in ws_current.iter_rows(min_row=2, values_only=False):
            app_value = row[app_col_curr - 1].value
            ctrl_value = row[ctrl_col_curr - 1].value
            key = (app_value, ctrl_value)
            if app_value and ctrl_value:
                current_data[key] = row

        for key, previous_row in previous_data.items():
            current_row = current_data.get(key)
            if current_row:
                for column, (col_idx_prev, col_idx_curr) in columns.items():
                    previous_value = previous_row[col_idx_prev - 1].value
                    current_value = current_row[col_idx_curr - 1].value
                    cell_output = ws_current.cell(row=current_row[0].row, column=col_idx_curr)

                    if previous_value == current_value:
                        continue

                    try:
                        prev_value_num = float(previous_value)
                        curr_value_num = float(current_value)
                        formatted_prev_value = f"{prev_value_num:.2f}"
                        formatted_curr_value = f"{curr_value_num:.2f}"

                        if curr_value_num > prev_value_num:
                            cell_output.fill = green_fill
                            cell_output.value = f"{formatted_prev_value} → {formatted_curr_value} (Improved)"
                        else:
                            cell_output.fill = red_fill
                            cell_output.value = f"{formatted_prev_value} → {formatted_curr_value} (Declined)"
                    except ValueError:
                        logging.error(
                            f"Non-numeric value encountered for column '{column}': "
                            f"Previous={previous_value}, Current={current_value}"
                        )

        for key, current_row in current_data.items():
            if key not in previous_data:
                row_index = ws_current.max_row + 1
                for col_num, cell in enumerate(current_row, 1):
                    new_cell = ws_current.cell(row=row_index, column=col_num, value=cell.value)
                    new_cell.fill = added_fill

    except Exception as e:
        logging.error(f"Error in compare_machineagentsapm: {e}", exc_info=True)
        raise


def compare_datacollectorsapm(ws_previous, ws_current):
    try:
        columns = {
            'numberOfDataCollectorFieldsConfigured': None,
            'numberOfDataCollectorFieldsCollectedInSnapshots': None,
            'numberOfDataCollectorFieldsCollectedInAnalytics': None,
            'biqEnabled': None
        }

        for column in columns.keys():
            col_idx_prev = get_key_column(ws_previous, column)
            col_idx_curr = get_key_column(ws_current, column)
            if col_idx_prev is None or col_idx_curr is None:
                logging.error(f"The '{column}' column is missing in one of the sheets. Cannot proceed with comparison.")
                return
            columns[column] = (col_idx_prev, col_idx_curr)

        app_col_prev = get_key_column(ws_previous, 'application')
        app_col_curr = get_key_column(ws_current, 'application')
        ctrl_col_prev = get_key_column(ws_previous, 'controller')
        ctrl_col_curr = get_key_column(ws_current, 'controller')

        if app_col_prev is None or app_col_curr is None:
            logging.error("The 'application' column is missing in one of the sheets. Cannot proceed with comparison.")
            return

        if ctrl_col_prev is None or ctrl_col_curr is None:
            logging.error("The 'controller' column is missing in one of the sheets. This might affect the comparison.")
            return

        previous_data = {}
        current_data = {}

        for row in ws_previous.iter_rows(min_row=2, values_only=False):
            app_value = row[app_col_prev - 1].value
            ctrl_value = row[ctrl_col_prev - 1].value
            key = (app_value, ctrl_value)
            if app_value and ctrl_value:
                previous_data[key] = row

        for row in ws_current.iter_rows(min_row=2, values_only=False):
            app_value = row[app_col_curr - 1].value
            ctrl_value = row[ctrl_col_curr - 1].value
            key = (app_value, ctrl_value)
            if app_value and ctrl_value:
                current_data[key] = row

        for key, previous_row in previous_data.items():
            current_row = current_data.get(key)
            if current_row:
                for column, (col_idx_prev, col_idx_curr) in columns.items():
                    previous_value = previous_row[col_idx_prev - 1].value
                    current_value = current_row[col_idx_curr - 1].value
                    cell_output = ws_current.cell(row=current_row[0].row, column=col_idx_curr)

                    if previous_value == current_value:
                        cell_output.value = previous_value
                        continue

                    if column == 'biqEnabled':
                        prev_value_str = str(previous_value).strip().upper()
                        curr_value_str = str(current_value).strip().upper()

                        logging.info(f"Comparing {column}: Previous={prev_value_str}, Current={curr_value_str}")

                        if prev_value_str == "FALSE" and curr_value_str == "TRUE":
                            cell_output.fill = green_fill
                            cell_output.value = f"{previous_value} → {current_value} (Improved)"
                        elif prev_value_str == "TRUE" and curr_value_str == "FALSE":
                            cell_output.fill = red_fill
                            cell_output.value = f"{previous_value} → {current_value} (Declined)"
                        else:
                            cell_output.fill = red_fill
                            cell_output.value = f"{previous_value} → {current_value} (Changed)"

        for key, current_row in current_data.items():
            if key not in previous_data:
                row_index = ws_current.max_row + 1
                for col_num, cell in enumerate(current_row, 1):
                    new_cell = ws_current.cell(row=row_index, column=col_num, value=cell.value)
                    new_cell.fill = added_fill

    except Exception as e:
        logging.error(f"Error in compare_datacollectorsapm: {e}", exc_info=True)
        raise


def compare_backendsapm(ws_previous, ws_current):
    try:
        columns = {
            'percentBackendsWithLoad': None,
            'backendLimitNotHit': None,
            'numberOfCustomBackendRules': None
        }

        for column in columns.keys():
            col_idx_prev = get_key_column(ws_previous, column)
            col_idx_curr = get_key_column(ws_current, column)
            if col_idx_prev is None or col_idx_curr is None:
                logging.error(f"The '{column}' column is missing in one of the sheets. Cannot proceed with comparison.")
                return
            columns[column] = (col_idx_prev, col_idx_curr)

        app_col_prev = get_key_column(ws_previous, 'application')
        app_col_curr = get_key_column(ws_current, 'application')
        ctrl_col_prev = get_key_column(ws_previous, 'controller')
        ctrl_col_curr = get_key_column(ws_current, 'controller')

        if app_col_prev is None or app_col_curr is None:
            logging.error("The 'application' column is missing in one of the sheets. Cannot proceed with comparison.")
            return

        if ctrl_col_prev is None or ctrl_col_curr is None:
            logging.error("The 'controller' column is missing in one of the sheets. This might affect the comparison.")
            return

        previous_data = {}
        current_data = {}

        for row in ws_previous.iter_rows(min_row=2, values_only=False):
            app_value = row[app_col_prev - 1].value
            ctrl_value = row[ctrl_col_prev - 1].value
            key = (app_value, ctrl_value)
            if app_value and ctrl_value:
                previous_data[key] = row

        for row in ws_current.iter_rows(min_row=2, values_only=False):
            app_value = row[app_col_curr - 1].value
            ctrl_value = row[ctrl_col_curr - 1].value
            key = (app_value, ctrl_value)
            if app_value and ctrl_value:
                current_data[key] = row

        for key, previous_row in previous_data.items():
            current_row = current_data.get(key)
            if current_row:
                for column, (col_idx_prev, col_idx_curr) in columns.items():
                    previous_value = previous_row[col_idx_prev - 1].value
                    current_value = current_row[col_idx_curr - 1].value
                    cell_output = ws_current.cell(row=current_row[0].row, column=col_idx_curr)

                    if previous_value == current_value:
                        cell_output.value = previous_value
                        continue

                    if column == 'backendLimitNotHit':
                        prev_value = str(previous_value).strip().upper()
                        curr_value = str(current_value).strip().upper()

                        logging.info(f"Comparing backendLimitNotHit: Previous={prev_value}, Current={curr_value}")

                        if prev_value == "FALSE" and curr_value == "TRUE":
                            cell_output.fill = green_fill
                            cell_output.value = "FALSE → TRUE"
                        elif prev_value == "TRUE" and curr_value == "FALSE":
                            cell_output.fill = red_fill
                            cell_output.value = "TRUE → FALSE"
                        else:
                            cell_output.fill = red_fill
                            cell_output.value = f"{previous_value} → {current_value} (Changed)"
                    else:
                        try:
                            prev_value_num = float(previous_value)
                            curr_value_num = float(current_value)
                            formatted_prev_value = f"{prev_value_num:.2f}"
                            formatted_curr_value = f"{curr_value_num:.2f}"

                            if curr_value_num > prev_value_num:
                                cell_output.fill = green_fill
                                cell_output.value = f"{formatted_prev_value} → {formatted_curr_value} (Increased)"
                            else:
                                cell_output.fill = red_fill
                                cell_output.value = f"{formatted_prev_value} → {formatted_curr_value} (Decreased)"
                        except ValueError:
                            logging.error(
                                f"Non-numeric value encountered for column '{column}': "
                                f"Previous={previous_value}, Current={current_value}"
                            )

        for key, current_row in current_data.items():
            if key not in previous_data:
                row_index = ws_current.max_row + 1
                for col_num, cell in enumerate(current_row, 1):
                    new_cell = ws_current.cell(row=row_index, column=col_num, value=cell.value)
                    new_cell.fill = added_fill

    except Exception as e:
        logging.error(f"Error in compare_backendsapm: {e}", exc_info=True)
        raise


def compare_overheadapm(ws_previous, ws_current):
    try:
        columns = {
            'developerModeNotEnabledForAnyBT': None,
            'findEntryPointsNotEnabled': None,
            'aggressiveSnapshottingNotEnabled': None,
            'developerModeNotEnabledForApplication': None
        }

        for column in columns.keys():
            col_idx_prev = get_key_column(ws_previous, column)
            col_idx_curr = get_key_column(ws_current, column)
            if col_idx_prev is None or col_idx_curr is None:
                logging.error(f"The '{column}' column is missing in one of the sheets. Cannot proceed with comparison.")
                return
            columns[column] = (col_idx_prev, col_idx_curr)

        app_col_prev = get_key_column(ws_previous, 'application')
        app_col_curr = get_key_column(ws_current, 'application')
        ctrl_col_prev = get_key_column(ws_previous, 'controller')
        ctrl_col_curr = get_key_column(ws_current, 'controller')

        if app_col_prev is None or app_col_curr is None:
            logging.error("The 'application' column is missing in one of the sheets. Cannot proceed with comparison.")
            return

        if ctrl_col_prev is None or ctrl_col_curr is None:
            logging.error("The 'controller' column is missing in one of the sheets. This might affect the comparison.")
            return

        previous_data = {}
        current_data = {}

        for row in ws_previous.iter_rows(min_row=2, values_only=False):
            app_value = row[app_col_prev - 1].value
            ctrl_value = row[ctrl_col_prev - 1].value
            key = (app_value, ctrl_value)
            if app_value and ctrl_value:
                previous_data[key] = row

        for row in ws_current.iter_rows(min_row=2, values_only=False):
            app_value = row[app_col_curr - 1].value
            ctrl_value = row[ctrl_col_curr - 1].value
            key = (app_value, ctrl_value)
            if app_value and ctrl_value:
                current_data[key] = row

        for key, previous_row in previous_data.items():
            current_row = current_data.get(key)
            if current_row:
                for column, (col_idx_prev, col_idx_curr) in columns.items():
                    previous_value = previous_row[col_idx_prev - 1].value
                    current_value = current_row[col_idx_curr - 1].value
                    cell_output = ws_current.cell(row=current_row[0].row, column=col_idx_curr)

                    if previous_value == current_value:
                        continue

                    if column in (
                        'developerModeNotEnabledForAnyBT',
                        'findEntryPointsNotEnabled',
                        'aggressiveSnapshottingNotEnabled',
                        'developerModeNotEnabledForApplication',
                    ):
                        prev_value_str = str(previous_value).strip().upper()
                        curr_value_str = str(current_value).strip().upper()

                        logging.info(f"Comparing {column}: Previous={prev_value_str}, Current={curr_value_str}")

                        if prev_value_str == "FALSE" and curr_value_str == "TRUE":
                            cell_output.fill = green_fill
                            cell_output.value = f"{previous_value} → {current_value} (Improved)"
                        elif prev_value_str == "TRUE" and curr_value_str == "FALSE":
                            cell_output.fill = red_fill
                            cell_output.value = f"{previous_value} → {current_value} (Declined)"
                        else:
                            cell_output.fill = red_fill
                            cell_output.value = f"{previous_value} → {current_value} (Changed)"

        for key, current_row in current_data.items():
            if key not in previous_data:
                row_index = ws_current.max_row + 1
                for col_num, cell in enumerate(current_row, 1):
                    new_cell = ws_current.cell(row=row_index, column=col_num, value=cell.value)
                    new_cell.fill = added_fill

    except Exception as e:
        logging.error(f"Error in compare_overheadapm: {e}", exc_info=True)
        raise


def compare_healthrulesandalertingapm(ws_previous, ws_current):
    try:
        columns = {
            'numberOfHealthRuleViolations': None,
            'numberOfDefaultHealthRulesModified': None,
            'numberOfActionsBoundToEnabledPolicies': None,
            'numberOfCustomHealthRules': None
        }

        for column in columns.keys():
            col_idx_prev = get_key_column(ws_previous, column)
            col_idx_curr = get_key_column(ws_current, column)
            if col_idx_prev is None or col_idx_curr is None:
                logging.error(f"The '{column}' column is missing in one of the sheets. Cannot proceed with comparison.")
                return
            columns[column] = (col_idx_prev, col_idx_curr)

        app_col_prev = get_key_column(ws_previous, 'application')
        app_col_curr = get_key_column(ws_current, 'application')
        ctrl_col_prev = get_key_column(ws_previous, 'controller')
        ctrl_col_curr = get_key_column(ws_current, 'controller')

        if app_col_prev is None or app_col_curr is None:
            logging.error("The 'application' column is missing in one of the sheets. Cannot proceed with comparison.")
            return

        if ctrl_col_prev is None or ctrl_col_curr is None:
            logging.error("The 'controller' column is missing in one of the sheets. This might affect the comparison.")
            return

        previous_data = {}
        current_data = {}

        for row in ws_previous.iter_rows(min_row=2, values_only=False):
            app_value = row[app_col_prev - 1].value
            ctrl_value = row[ctrl_col_prev - 1].value
            key = (app_value, ctrl_value)
            if app_value and ctrl_value:
                previous_data[key] = row

        for row in ws_current.iter_rows(min_row=2, values_only=False):
            app_value = row[app_col_curr - 1].value
            ctrl_value = row[ctrl_col_curr - 1].value
            key = (app_value, ctrl_value)
            if app_value and ctrl_value:
                current_data[key] = row

        for key, previous_row in previous_data.items():
            current_row = current_data.get(key)
            if current_row:
                for column, (col_idx_prev, col_idx_curr) in columns.items():
                    previous_value = previous_row[col_idx_prev - 1].value
                    current_value = current_row[col_idx_curr - 1].value
                    cell_output = ws_current.cell(row=current_row[0].row, column=col_idx_curr)

                    if previous_value == current_value:
                        continue

                    try:
                        prev_value_num = float(previous_value)
                        curr_value_num = float(current_value)
                        formatted_prev_value = f"{prev_value_num:.2f}"
                        formatted_curr_value = f"{curr_value_num:.2f}"

                        if column == 'numberOfHealthRuleViolations':
                            if curr_value_num > prev_value_num:
                                cell_output.fill = red_fill
                                cell_output.value = f"{formatted_prev_value} → {formatted_curr_value} (Increased)"
                            else:
                                cell_output.fill = green_fill
                                cell_output.value = f"{formatted_prev_value} → {formatted_curr_value} (Decreased)"
                        else:
                            if curr_value_num > prev_value_num:
                                cell_output.fill = green_fill
                                cell_output.value = f"{formatted_prev_value} → {formatted_curr_value} (Increased)"
                            else:
                                cell_output.fill = red_fill
                                cell_output.value = f"{formatted_prev_value} → {formatted_curr_value} (Decreased)"
                    except ValueError:
                        logging.error(
                            f"Non-numeric value encountered for column '{column}': "
                            f"Previous={previous_value}, Current={current_value}"
                        )

        for key, current_row in current_data.items():
            if key not in previous_data:
                row_index = ws_current.max_row + 1
                for col_num, cell in enumerate(current_row, 1):
                    new_cell = ws_current.cell(row=row_index, column=col_num, value=cell.value)
                    new_cell.fill = added_fill

    except Exception as e:
        logging.error(f"Error in compare_healthrulesandalertingapm: {e}", exc_info=True)
        raise


def compare_errorconfigurationapm(ws_previous, ws_current):
    try:
        columns = {
            'successPercentageOfWorstTransaction': None,
            'numberOfCustomRules': None
        }

        for column in columns.keys():
            col_idx_prev = get_key_column(ws_previous, column)
            col_idx_curr = get_key_column(ws_current, column)
            if col_idx_prev is None or col_idx_curr is None:
                logging.error(f"The '{column}' column is missing in one of the sheets. Cannot proceed with comparison.")
                return
            columns[column] = (col_idx_prev, col_idx_curr)

        ctrl_col_prev = get_key_column(ws_previous, 'controller')
        ctrl_col_curr = get_key_column(ws_current, 'controller')
        app_col_prev = get_key_column(ws_previous, 'application')
        app_col_curr = get_key_column(ws_current, 'application')

        if app_col_prev is None or app_col_curr is None:
            logging.error("The 'application' column is missing in one of the sheets. Cannot proceed with comparison.")
            return

        if ctrl_col_prev is None or ctrl_col_curr is None:
            logging.error("The 'controller' column is missing in one of the sheets. This might affect the comparison.")
            return

        previous_data = {}
        current_data = {}

        for row in ws_previous.iter_rows(min_row=2, values_only=False):
            app_value = row[app_col_prev - 1].value
            ctrl_value = row[ctrl_col_prev - 1].value
            key = (app_value, ctrl_value)
            if app_value and ctrl_value:
                previous_data[key] = row

        for row in ws_current.iter_rows(min_row=2, values_only=False):
            app_value = row[app_col_curr - 1].value
            ctrl_value = row[ctrl_col_curr - 1].value
            key = (app_value, ctrl_value)
            if app_value and ctrl_value:
                current_data[key] = row

        for key, previous_row in previous_data.items():
            current_row = current_data.get(key)
            if current_row:
                for column, (col_idx_prev, col_idx_curr) in columns.items():
                    previous_value = previous_row[col_idx_prev - 1].value
                    current_value = current_row[col_idx_curr - 1].value
                    cell_output = ws_current.cell(row=current_row[0].row, column=col_idx_curr)

                    if previous_value == current_value:
                        continue

                    try:
                        prev_value_num = float(previous_value)
                        curr_value_num = float(current_value)
                        formatted_prev_value = f"{prev_value_num:.2f}"
                        formatted_curr_value = f"{curr_value_num:.2f}"

                        if column in ('successPercentageOfWorstTransaction', 'numberOfCustomRules'):
                            if curr_value_num > prev_value_num:
                                cell_output.fill = green_fill
                                cell_output.value = f"{formatted_prev_value} → {formatted_curr_value} (Increased)"
                            else:
                                cell_output.fill = red_fill
                                cell_output.value = f"{formatted_prev_value} → {formatted_curr_value} (Decreased)"
                    except ValueError:
                        logging.error(
                            f"Non-numeric value encountered for column '{column}': "
                            f"Previous={previous_value}, Current={current_value}"
                        )

        for key, current_row in current_data.items():
            if key not in previous_data:
                row_index = ws_current.max_row + 1
                for col_num, cell in enumerate(current_row, 1):
                    new_cell = ws_current.cell(row=row_index, column=col_num, value=cell.value)
                    new_cell.fill = added_fill

    except Exception as e:
        logging.error(f"Error in compare_errorconfigurationapm: {e}", exc_info=True)
        raise


def compare_serviceendpointsapm(ws_previous, ws_current):
    try:
        columns = {
            'numberOfCustomServiceEndpointRules': None,
            'serviceEndpointLimitNotHit': None,
            'percentServiceEndpointsWithLoadOrDisabled': None
        }

        for column in columns.keys():
            col_idx_prev = get_key_column(ws_previous, column)
            col_idx_curr = get_key_column(ws_current, column)
            if col_idx_prev is None or col_idx_curr is None:
                logging.error(f"The '{column}' column is missing in one of the sheets. Cannot proceed with comparison.")
                return
            columns[column] = (col_idx_prev, col_idx_curr)

        ctrl_col_prev = get_key_column(ws_previous, 'controller')
        ctrl_col_curr = get_key_column(ws_current, 'controller')
        app_col_prev = get_key_column(ws_previous, 'application')
        app_col_curr = get_key_column(ws_current, 'application')

        if app_col_prev is None or app_col_curr is None:
            logging.error("The 'application' column is missing in one of the sheets. Cannot proceed with comparison.")
            return

        if ctrl_col_prev is None or ctrl_col_curr is None:
            logging.error("The 'controller' column is missing in one of the sheets. This might affect the comparison.")
            return

        previous_data = {}
        current_data = {}

        for row in ws_previous.iter_rows(min_row=2, values_only=False):
            app_value = row[app_col_prev - 1].value
            ctrl_value = row[ctrl_col_prev - 1].value
            key = (app_value, ctrl_value)
            if app_value and ctrl_value:
                previous_data[key] = row

        for row in ws_current.iter_rows(min_row=2, values_only=False):
            app_value = row[app_col_curr - 1].value
            ctrl_value = row[ctrl_col_curr - 1].value
            key = (app_value, ctrl_value)
            if app_value and ctrl_value:
                current_data[key] = row

        for key, previous_row in previous_data.items():
            current_row = current_data.get(key)
            if current_row:
                for column, (col_idx_prev, col_idx_curr) in columns.items():
                    previous_value = previous_row[col_idx_prev - 1].value
                    current_value = current_row[col_idx_curr - 1].value
                    cell_output = ws_current.cell(row=current_row[0].row, column=col_idx_curr)

                    logging.debug(
                        f"Comparing '{column}' for key '{key}': "
                        f"Previous={previous_value}, Current={current_value}"
                    )

                    if previous_value == current_value:
                        continue

                    try:
                        if column == 'numberOfCustomServiceEndpointRules':
                            prev_value_num = float(previous_value)
                            curr_value_num = float(current_value)
                            if curr_value_num > prev_value_num:
                                cell_output.fill = green_fill
                                cell_output.value = f"{prev_value_num:.2f} → {curr_value_num:.2f} (Increased)"
                            elif curr_value_num < prev_value_num:
                                cell_output.fill = red_fill
                                cell_output.value = f"{prev_value_num:.2f} → {curr_value_num:.2f} (Decreased)"

                        elif column == 'serviceEndpointLimitNotHit':
                            prev_value_str = str(previous_value).strip().upper()
                            curr_value_str = str(current_value).strip().upper()

                            logging.info(
                                f"Comparing serviceEndpointLimitNotHit: "
                                f"Previous={prev_value_str}, Current={curr_value_str}"
                            )

                            if prev_value_str == "FALSE" and curr_value_str == "TRUE":
                                cell_output.fill = green_fill
                                cell_output.value = "FALSE → TRUE"
                            elif prev_value_str == "TRUE" and curr_value_str == "FALSE":
                                cell_output.fill = red_fill
                                cell_output.value = "TRUE → FALSE"
                            else:
                                logging.info(
                                    f"Unexpected values for serviceEndpointLimitNotHit: "
                                    f"Previous={previous_value}, Current={current_value}"
                                )

                        elif column == 'percentServiceEndpointsWithLoadOrDisabled':
                            prev_value_num = float(previous_value)
                            curr_value_num = float(current_value)
                            if curr_value_num > prev_value_num:
                                cell_output.fill = green_fill
                                cell_output.value = f"{prev_value_num:.2f} → {curr_value_num:.2f} (Increased)"
                            elif curr_value_num < prev_value_num:
                                cell_output.fill = red_fill
                                cell_output.value = f"{prev_value_num:.2f} → {curr_value_num:.2f} (Decreased)"

                    except ValueError:
                        logging.error(
                            f"Non-numeric or invalid value encountered for column '{column}': "
                            f"Previous={previous_value}, Current={current_value}"
                        )

        for key, current_row in current_data.items():
            if key not in previous_data:
                row_index = ws_current.max_row + 1
                for col_num, cell in enumerate(current_row, 1):
                    new_cell = ws_current.cell(row=row_index, column=col_num, value=cell.value)
                    new_cell.fill = added_fill

    except Exception as e:
        logging.error(f"Error in compare_serviceendpointsapm: {e}", exc_info=True)
        raise


def compare_dashboardsapm(ws_previous, ws_current):
    try:
        columns = {
            'numberOfDashboards': None,
            'percentageOfDashboardsModifiedLast6Months': None,
            'numberOfDashboardsUsingBiQ': None
        }

        for column in columns.keys():
            col_idx_prev = get_key_column(ws_previous, column)
            col_idx_curr = get_key_column(ws_current, column)
            if col_idx_prev is None or col_idx_curr is None:
                logging.error(f"The '{column}' column is missing in one of the sheets. Cannot proceed with comparison.")
                return
            columns[column] = (col_idx_prev, col_idx_curr)

        ctrl_col_prev = get_key_column(ws_previous, 'controller')
        ctrl_col_curr = get_key_column(ws_current, 'controller')
        app_col_prev = get_key_column(ws_previous, 'application')
        app_col_curr = get_key_column(ws_current, 'application')

        if app_col_prev is None or app_col_curr is None:
            logging.error("The 'application' column is missing in one of the sheets. Cannot proceed with comparison.")
            return

        if ctrl_col_prev is None or ctrl_col_curr is None:
            logging.error("The 'controller' column is missing in one of the sheets. This might affect the comparison.")
            return

        previous_data = {}
        current_data = {}

        for row in ws_previous.iter_rows(min_row=2, values_only=False):
            app_value = row[app_col_prev - 1].value
            ctrl_value = row[ctrl_col_prev - 1].value
            key = (app_value, ctrl_value)
            if app_value and ctrl_value:
                previous_data[key] = row

        for row in ws_current.iter_rows(min_row=2, values_only=False):
            app_value = row[app_col_curr - 1].value
            ctrl_value = row[ctrl_col_curr - 1].value
            key = (app_value, ctrl_value)
            if app_value and ctrl_value:
                current_data[key] = row

        for key, previous_row in previous_data.items():
            current_row = current_data.get(key)
            if current_row:
                for column, (col_idx_prev, col_idx_curr) in columns.items():
                    previous_value = previous_row[col_idx_prev - 1].value
                    current_value = current_row[col_idx_curr - 1].value
                    cell_output = ws_current.cell(row=current_row[0].row, column=col_idx_curr)

                    if previous_value == current_value:
                        continue

                    try:
                        prev_value_num = float(previous_value)
                        curr_value_num = float(current_value)
                        if curr_value_num > prev_value_num:
                            cell_output.fill = green_fill
                            cell_output.value = f"{prev_value_num:.2f} → {curr_value_num:.2f} (Increased)"
                        elif curr_value_num < prev_value_num:
                            cell_output.fill = red_fill
                            cell_output.value = f"{prev_value_num:.2f} → {curr_value_num:.2f} (Decreased)"
                    except ValueError:
                        logging.error(
                            f"Non-numeric or invalid value encountered for column '{column}': "
                            f"Previous={previous_value}, Current={current_value}"
                        )

        for key, current_row in current_data.items():
            if key not in previous_data:
                row_index = ws_current.max_row + 1
                for col_num, cell in enumerate(current_row, 1):
                    new_cell = ws_current.cell(row=row_index, column=col_num, value=cell.value)
                    new_cell.fill = added_fill

    except Exception as e:
        logging.error(f"Error in compare_dashboardsapm: {e}", exc_info=True)
        raise


def compare_overallassessmentapm(ws_previous, ws_current):
    try:
        columns = {
            'percentageTotalPlatinum': None,
            'percentageTotalGoldOrBetter': None,
            'percentageTotalSilverOrBetter': None
        }

        for column in columns.keys():
            col_idx_prev = get_key_column(ws_previous, column)
            col_idx_curr = get_key_column(ws_current, column)
            if col_idx_prev is None or col_idx_curr is None:
                logging.error(f"The '{column}' column is missing in one of the sheets. Cannot proceed with comparison.")
                return
            columns[column] = (col_idx_prev, col_idx_curr)

        ctrl_col_prev = get_key_column(ws_previous, 'controller')
        ctrl_col_curr = get_key_column(ws_current, 'controller')
        app_col_prev = get_key_column(ws_previous, 'application')
        app_col_curr = get_key_column(ws_current, 'application')

        if app_col_prev is None or app_col_curr is None:
            logging.error("The 'application' column is missing in one of the sheets. Cannot proceed with comparison.")
            return

        if ctrl_col_prev is None or ctrl_col_curr is None:
            logging.error("The 'controller' column is missing in one of the sheets. This might affect the comparison.")
            return

        previous_data = {}
        current_data = {}

        for row in ws_previous.iter_rows(min_row=2, values_only=False):
            app_value = row[app_col_prev - 1].value
            ctrl_value = row[ctrl_col_prev - 1].value
            key = (app_value, ctrl_value)
            if app_value and ctrl_value:
                previous_data[key] = row

        for row in ws_current.iter_rows(min_row=2, values_only=False):
            app_value = row[app_col_curr - 1].value
            ctrl_value = row[ctrl_col_curr - 1].value
            key = (app_value, ctrl_value)
            if app_value and ctrl_value:
                current_data[key] = row

        for key, previous_row in previous_data.items():
            current_row = current_data.get(key)
            if current_row:
                for column, (col_idx_prev, col_idx_curr) in columns.items():
                    previous_value = previous_row[col_idx_prev - 1].value
                    current_value = current_row[col_idx_curr - 1].value
                    cell_output = ws_current.cell(row=current_row[0].row, column=col_idx_curr)

                    if previous_value == current_value:
                        continue

                    try:
                        prev_value_num = float(previous_value)
                        curr_value_num = float(current_value)
                        if curr_value_num > prev_value_num:
                            cell_output.fill = green_fill
                            cell_output.value = f"{prev_value_num:.2f} → {curr_value_num:.2f} (Increased)"
                        elif curr_value_num < prev_value_num:
                            cell_output.fill = red_fill
                            cell_output.value = f"{prev_value_num:.2f} → {curr_value_num:.2f} (Decreased)"
                    except ValueError:
                        logging.error(
                            f"Non-numeric or invalid value encountered for column '{column}': "
                            f"Previous={previous_value}, Current={current_value}"
                        )

        for key, current_row in current_data.items():
            if key not in previous_data:
                row_index = ws_current.max_row + 1
                for col_num, cell in enumerate(current_row, 1):
                    new_cell = ws_current.cell(row=row_index, column=col_num, value=cell.value)
                    new_cell.fill = added_fill

    except Exception as e:
        logging.error(f"Error in compare_overallassessmentapm: {e}", exc_info=True)
        raise


def compare_businesstransactionsapm(ws_previous, ws_current):
    try:
        columns = {
            'numberOfBTs': None,
            'percentBTsWithLoad': None,
            'btLockdownEnabled': None,
            'numberCustomMatchRules': None
        }

        for column in columns.keys():
            col_idx_prev = get_key_column(ws_previous, column)
            col_idx_curr = get_key_column(ws_current, column)
            if col_idx_prev is None or col_idx_curr is None:
                print(f"The '{column}' column is missing in one of the sheets. Cannot proceed with comparison.")
                return
            columns[column] = (col_idx_prev, col_idx_curr)

        ctrl_col_prev = get_key_column(ws_previous, 'controller')
        ctrl_col_curr = get_key_column(ws_current, 'controller')
        app_col_prev = get_key_column(ws_previous, 'application')
        app_col_curr = get_key_column(ws_current, 'application')

        if app_col_prev is None or app_col_curr is None:
            print("The 'application' column is missing in one of the sheets. Cannot proceed with comparison.")
            return

        if ctrl_col_prev is None or ctrl_col_curr is None:
            print("The 'controller' column is missing in one of the sheets. This might affect the comparison.")
            return

        previous_data = {}
        current_data = {}

        for row in ws_previous.iter_rows(min_row=2, values_only=False):
            app_value = row[app_col_prev - 1].value
            ctrl_value = row[ctrl_col_prev - 1].value
            key = (app_value, ctrl_value)
            if app_value and ctrl_value:
                previous_data[key] = row

        for row in ws_current.iter_rows(min_row=2, values_only=False):
            app_value = row[app_col_curr - 1].value
            ctrl_value = row[ctrl_col_curr - 1].value
            key = (app_value, ctrl_value)
            if app_value and ctrl_value:
                current_data[key] = row

        for key, previous_row in previous_data.items():
            current_row = current_data.get(key)
            if current_row:
                for column, (col_idx_prev, col_idx_curr) in columns.items():
                    previous_value = previous_row[col_idx_prev - 1].value
                    current_value = current_row[col_idx_curr - 1].value
                    cell_output = ws_current.cell(row=current_row[0].row, column=col_idx_curr)

                    if previous_value == current_value:
                        continue

                    try:     

                        if column == 'numberOfBTs':
                            prev_value_num = float(previous_value)
                            curr_value_num = float(current_value)
                            if 201 <= prev_value_num <= 600:
                                if curr_value_num < prev_value_num:
                                    cell_output.fill = green_fill
                                    cell_output.value = (
                                        f"{prev_value_num:.2f} → {curr_value_num:.2f} (Decreased)"
                                    )
                                elif curr_value_num > prev_value_num:
                                    cell_output.fill = red_fill
                                    cell_output.value = (
                                        f"{prev_value_num:.2f} → {curr_value_num:.2f} (Increased)"
                                    )

                        elif column == 'percentBTsWithLoad':
                            prev_value_num = float(previous_value)
                            curr_value_num = float(current_value)
                            if curr_value_num > prev_value_num:
                                cell_output.fill = green_fill
                                cell_output.value = (
                                    f"{prev_value_num:.2f} → {curr_value_num:.2f} (Increased)"
                                )
                            elif curr_value_num < prev_value_num:
                                cell_output.fill = red_fill
                                cell_output.value = (
                                    f"{prev_value_num:.2f} → {curr_value_num:.2f} (Decreased)"
                                )

                        elif column == 'btLockdownEnabled':
                            prev_value_str = str(previous_value).strip().upper()
                            curr_value_str = str(current_value).strip().upper()

                            print(
                                f"Comparing btLockdownEnabled for app {key}: "
                                f"Previous={prev_value_str}, Current={curr_value_str}"
                            )

                            if prev_value_str == "FALSE" and curr_value_str == "TRUE":
                                print(f"Update: FALSE → TRUE for app {key}")
                                cell_output.fill = green_fill
                                cell_output.value = "FALSE → TRUE"
                            elif prev_value_str == "TRUE" and curr_value_str == "FALSE":
                                print(f"Update: TRUE → FALSE for app {key}")
                                cell_output.fill = red_fill
                                cell_output.value = "TRUE → FALSE"
                            else:
                                print(
                                    f"Unexpected values for btLockdownEnabled: "
                                    f"Previous={previous_value}, Current={current_value}"
                                )

                        elif column == 'numberCustomMatchRules':
                            prev_value_num = float(previous_value)
                            curr_value_num = float(current_value)
                            if curr_value_num > prev_value_num:
                                cell_output.fill = green_fill
                                cell_output.value = (
                                    f"{prev_value_num:.2f} → {curr_value_num:.2f} (Increased)"
                                )
                            elif curr_value_num < prev_value_num:
                                cell_output.fill = red_fill
                                cell_output.value = (
                                    f"{prev_value_num:.2f} → {curr_value_num:.2f} (Decreased)"
                                )
                    except ValueError:
                        print(
                            f"Non-numeric or invalid value encountered for column '{column}': "
                            f"Previous={previous_value}, Current={current_value}"
                        )

        for key, current_row in current_data.items():
            if key not in previous_data:
                row_index = ws_current.max_row + 1
                for col_num, cell in enumerate(current_row, 1):
                    new_cell = ws_current.cell(row=row_index, column=col_num, value=cell.value)
                    new_cell.fill = added_fill

    except Exception as e:
        print(f"Error in compare_businesstransactionsapm: {e}")
        raise

# Map APM sheet names → their comparer functions
SHEET_COMPARERS_APM = {
    "Analysis": compare_analysis,
    "AppAgentsAPM": compare_appagentsapm,
    "MachineAgentsAPM": compare_machineagentsapm,
    "DataCollectorsAPM": compare_datacollectorsapm,
    "BackendsAPM": compare_backendsapm,
    "OverheadAPM": compare_overheadapm,
    "HealthRulesAndAlertingAPM": compare_healthrulesandalertingapm,
    "ErrorConfigurationAPM": compare_errorconfigurationapm,
    "ServiceEndpointsAPM": compare_serviceendpointsapm,
    "DashboardsAPM": compare_dashboardsapm,
    "OverallAssessmentAPM": compare_overallassessmentapm,
    "BusinessTransactionsAPM": compare_businesstransactionsapm,
}


def compare_files_other_sheets_apm(
    previous_file_path: str,
    current_file_path: str,
    output_file_path: str,
) -> None:
    """
    APM-only sheet dispatcher (this is what compare_tool.comparers imports).
    """
    try:
        wb_previous = load_workbook(previous_file_path)
        wb_current = load_workbook(current_file_path)

        for sheet_name in wb_current.sheetnames:
            if sheet_name == "Summary":
                # Summary is handled elsewhere (copied only)
                continue

            if sheet_name not in wb_previous.sheetnames:
                logger.warning("APM sheet '%s' missing in previous workbook.", sheet_name)
                continue

            ws_prev = wb_previous[sheet_name]
            ws_curr = wb_current[sheet_name]

            logger.debug("Processing APM sheet: %s", sheet_name)

            compare_func = SHEET_COMPARERS_APM.get(sheet_name)
            if compare_func:
                compare_func(ws_prev, ws_curr)
            else:
                logger.warning("No comparer defined for APM sheet: %s", sheet_name)

        wb_current.save(output_file_path)
        logger.info("APM comparison results saved to: %s", output_file_path)

    except Exception as e:
        logger.error("Error in compare_files_other_sheets_apm: %s", e, exc_info=True)
        raise
