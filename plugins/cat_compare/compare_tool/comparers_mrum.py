"""
comparers_mrum.py
-----------------
This module contains the logic for comparing MRUM (Metric Rule Update Management) data.

Purpose:
- Compares data between the "previous" and "current" MRUM Excel files.
- Handles sheet-specific comparisons using defined comparer functions.
- Logs warnings for missing sheets or undefined comparers.

Key Features:
- Iterates through all sheets in the "current" workbook.
- Uses `SHEET_COMPARERS_MRUM` to apply sheet-specific comparison logic.
- Saves the comparison results to the specified output file.
- Logs detailed information about the comparison process.

Key Functions:
- `compare_files_other_sheets_mrum`: Compares all sheets except "Summary".
"""

# compare_tool/comparers_mrum.py

import logging
from typing import Dict, Tuple, Any

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from .excel_io import get_key_column

logger = logging.getLogger(__name__)

# Color fills (copy same as APM/BRUM)
red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
added_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')


# =========================
# MRUM COMPARISON FUNCTIONS
# =========================

def compare_analysis_mrum(ws_previous: Worksheet, ws_current: Worksheet) -> None:
    """
    Compare 'Analysis' sheet for MRUM.
    Columns:
      - NetworkRequestsMRUM
      - HealthRulesAndAlertingMRUM
      - OverallAssessment
    Key columns:
      - name
      - controller
    """
    try:
        columns = {
            'NetworkRequestsMRUM': None,
            'HealthRulesAndAlertingMRUM': None,
            'OverallAssessment': None,
        }

        # Locate comparison columns in both previous and current
        for column in columns.keys():
            col_idx_prev = get_key_column(ws_previous, column)
            col_idx_curr = get_key_column(ws_current, column)
            if col_idx_prev is None or col_idx_curr is None:
                logger.error("[MRUM] '%s' missing in Analysis.", column)
                return
            columns[column] = (col_idx_prev, col_idx_curr)

        # Key columns
        ctrl_col_prev = get_key_column(ws_previous, 'controller')
        ctrl_col_curr = get_key_column(ws_current, 'controller')
        name_col_prev = get_key_column(ws_previous, 'name')
        name_col_curr = get_key_column(ws_current, 'name')

        if name_col_prev is None or name_col_curr is None:
            logger.error("[MRUM] 'name' missing in Analysis.")
            return

        previous_data = {}
        current_data = {}

        # Build maps keyed by (name, controller)
        for row in ws_previous.iter_rows(min_row=2, values_only=False):
            key = (
                row[name_col_prev - 1].value,
                row[ctrl_col_prev - 1].value if ctrl_col_prev else None,
            )
            if key[0]:
                previous_data[key] = row

        for row in ws_current.iter_rows(min_row=2, values_only=False):
            key = (
                row[name_col_curr - 1].value,
                row[ctrl_col_curr - 1].value if ctrl_col_curr else None,
            )
            if key[0]:
                current_data[key] = row

        ranking = {'bronze': 1, 'silver': 2, 'gold': 3, 'platinum': 4}

        # Compare per key
        for key, prev_row in previous_data.items():
            cur_row = current_data.get(key)
            if not cur_row:
                continue

            for column, (p_idx, c_idx) in columns.items():
                prev_val = prev_row[p_idx - 1].value
                cur_val = cur_row[c_idx - 1].value
                if prev_val == cur_val:
                    continue

                cell_out = ws_current.cell(row=cur_row[0].row, column=c_idx)
                prev_rank = ranking.get(str(prev_val).strip().lower(), 0)
                cur_rank = ranking.get(str(cur_val).strip().lower(), 0)

                if cur_rank > prev_rank:
                    cell_out.fill = green_fill
                    cell_out.value = f"{prev_val} → {cur_val} (Upgraded)"
                elif cur_rank < prev_rank:
                    cell_out.fill = red_fill
                    cell_out.value = f"{prev_val} → {cur_val} (Downgraded)"

        # Add new entries not present in previous
        for key, cur_row in current_data.items():
            if key not in previous_data:
                row_index = ws_current.max_row + 1
                for col_num, cell in enumerate(cur_row, 1):
                    new_cell = ws_current.cell(
                        row=row_index,
                        column=col_num,
                        value=cell.value,
                    )
                    new_cell.fill = added_fill

    except Exception as e:
        logger.error("[MRUM] Error in compare_analysis_mrum: %s", e, exc_info=True)
        raise


def compare_networkrequestsmrum(ws_previous: Worksheet, ws_current: Worksheet) -> None:
    """
    Compare 'NetworkRequestsMRUM' sheet.
    """
    try:
        columns = {
            'collectingDataPastOneDay': 'bool',
            'networkRequestLimitNotHit': 'bool',
            'numberCustomMatchRules': 'num',
            'hasBtCorrelation': 'bool',
            'hasCustomEventServiceIncludeRule': 'bool',
        }

        idx: Dict[str, Tuple[int, int]] = {}
        for c in columns:
            p = get_key_column(ws_previous, c)
            n = get_key_column(ws_current, c)
            if p is None or n is None:
                logger.error("[MRUM] Missing '%s' in NetworkRequestsMRUM.", c)
                return
            idx[c] = (p, n)

        app_prev = get_key_column(ws_previous, 'application')
        app_curr = get_key_column(ws_current, 'application')
        ctrl_prev = get_key_column(ws_previous, 'controller')
        ctrl_curr = get_key_column(ws_current, 'controller')

        previous_map = {}
        current_map = {}

        for row in ws_previous.iter_rows(min_row=2, values_only=False):
            key = (
                row[app_prev - 1].value,
                row[ctrl_prev - 1].value if ctrl_prev else None,
            )
            if key[0]:
                previous_map[key] = row

        for row in ws_current.iter_rows(min_row=2, values_only=False):
            key = (
                row[app_curr - 1].value,
                row[ctrl_curr - 1].value if ctrl_curr else None,
            )
            if key[0]:
                current_map[key] = row

        for key, prow in previous_map.items():
            crow = current_map.get(key)
            if not crow:
                continue

            for col, kind in columns.items():
                p_idx, c_idx = idx[col]
                pv = prow[p_idx - 1].value
                cv = crow[c_idx - 1].value
                if pv == cv:
                    continue

                cell = ws_current.cell(row=crow[0].row, column=c_idx)

                if kind == 'bool':
                    p_str = str(pv).strip().upper()
                    c_str = str(cv).strip().upper()

                    if p_str == 'FALSE' and c_str == 'TRUE':
                        cell.fill = green_fill
                        cell.value = f"{pv} → {cv} (Improved)"
                    elif p_str == 'TRUE' and c_str == 'FALSE':
                        cell.fill = red_fill
                        cell.value = f"{pv} → {cv} (Declined)"
                    else:
                        cell.fill = red_fill
                        cell.value = f"{pv} → {cv} (Changed)"
                else:
                    try:
                        pnum = float(pv)
                        cnum = float(cv)
                        if cnum > pnum:
                            cell.fill = green_fill
                            cell.value = f"{pnum:.2f} → {cnum:.2f} (Increased)"
                        else:
                            cell.fill = red_fill
                            cell.value = f"{pnum:.2f} → {cnum:.2f} (Decreased)"
                    except Exception:
                        logger.error(
                            "[MRUM] Non-numeric '%s': %s vs %s", col, pv, cv
                        )

        # New entries
        for key, crow in current_map.items():
            if key not in previous_map:
                r = ws_current.max_row + 1
                for i, c in enumerate(crow, 1):
                    nc = ws_current.cell(row=r, column=i, value=c.value)
                    nc.fill = added_fill

    except Exception as e:
        logger.error(
            "[MRUM] Error in compare_networkrequestsmrum: %s", e, exc_info=True
        )
        raise


def compare_healthrulesandalertingmrum(ws_previous: Worksheet, ws_current: Worksheet) -> None:
    """
    Compare 'HealthRulesAndAlertingMRUM' sheet.
    """
    try:
        columns = {
            'numberOfHealthRuleViolations': 'lower_better',
            'numberOfActionsBoundToEnabledPolicies': 'higher_better',
            'numberOfCustomHealthRules': 'higher_better',
        }

        idx: Dict[str, Tuple[int, int]] = {}
        for c in columns:
            p = get_key_column(ws_previous, c)
            n = get_key_column(ws_current, c)
            if p is None or n is None:
                logger.error("[MRUM] Missing '%s' in HealthRulesAndAlertingMRUM.", c)
                return
            idx[c] = (p, n)

        app_prev = get_key_column(ws_previous, 'application')
        app_curr = get_key_column(ws_current, 'application')
        ctrl_prev = get_key_column(ws_previous, 'controller')
        ctrl_curr = get_key_column(ws_current, 'controller')

        previous_map = {}
        current_map = {}

        for row in ws_previous.iter_rows(min_row=2, values_only=False):
            key = (
                row[app_prev - 1].value,
                row[ctrl_prev - 1].value if ctrl_prev else None,
            )
            if key[0]:
                previous_map[key] = row

        for row in ws_current.iter_rows(min_row=2, values_only=False):
            key = (
                row[app_curr - 1].value,
                row[ctrl_curr - 1].value if ctrl_curr else None,
            )
            if key[0]:
                current_map[key] = row

        for key, prow in previous_map.items():
            crow = current_map.get(key)
            if not crow:
                continue

            for col, rule in columns.items():
                p_idx, c_idx = idx[col]
                pv = prow[p_idx - 1].value
                cv = crow[c_idx - 1].value
                if pv == cv:
                    continue

                cell = ws_current.cell(row=crow[0].row, column=c_idx)

                try:
                    pnum = float(pv)
                    cnum = float(cv)

                    if rule == 'lower_better':
                        if cnum < pnum:
                            cell.fill = green_fill
                            cell.value = f"{pnum:.2f} → {cnum:.2f} (Improved)"
                        else:
                            cell.fill = red_fill
                            cell.value = f"{pnum:.2f} → {cnum:.2f} (Declined)"
                    else:
                        if cnum > pnum:
                            cell.fill = green_fill
                            cell.value = f"{pnum:.2f} → {cnum:.2f} (Increased)"
                        else:
                            cell.fill = red_fill
                            cell.value = f"{pnum:.2f} → {cnum:.2f} (Decreased)"
                except Exception:
                    logger.error(
                        "[MRUM] Non-numeric '%s': %s vs %s", col, pv, cv
                    )

        # New entries
        for key, crow in current_map.items():
            if key not in previous_map:
                r = ws_current.max_row + 1
                for i, c in enumerate(crow, 1):
                    nc = ws_current.cell(row=r, column=i, value=c.value)
                    nc.fill = added_fill

    except Exception as e:
        logger.error(
            "[MRUM] Error in compare_healthrulesandalertingmrum: %s", e, exc_info=True
        )
        raise


def compare_overallassessmentmrum(ws_previous: Worksheet, ws_current: Worksheet) -> None:
    """
    Compare 'OverallAssessmentMRUM' sheet.
    """
    try:
        columns = {
            'percentageTotalPlatinum': None,
            'percentageTotalGoldOrBetter': None,
            'percentageTotalSilverOrBetter': None,
        }

        for c in columns.keys():
            p = get_key_column(ws_previous, c)
            n = get_key_column(ws_current, c)
            if p is None or n is None:
                logger.error("[MRUM] Missing '%s' in OverallAssessmentMRUM.", c)
                return
            columns[c] = (p, n)

        ctrl_prev = get_key_column(ws_previous, 'controller')
        ctrl_curr = get_key_column(ws_current, 'controller')
        app_prev = get_key_column(ws_previous, 'application')
        app_curr = get_key_column(ws_current, 'application')

        previous_map = {}
        current_map = {}

        for row in ws_previous.iter_rows(min_row=2, values_only=False):
            key = (
                row[app_prev - 1].value,
                row[ctrl_prev - 1].value if ctrl_prev else None,
            )
            if key[0]:
                previous_map[key] = row

        for row in ws_current.iter_rows(min_row=2, values_only=False):
            key = (
                row[app_curr - 1].value,
                row[ctrl_curr - 1].value if ctrl_curr else None,
            )
            if key[0]:
                current_map[key] = row

        for key, prow in previous_map.items():
            crow = current_map.get(key)
            if not crow:
                continue

            for col, (p_idx, c_idx) in columns.items():
                pv = prow[p_idx - 1].value
                cv = crow[c_idx - 1].value
                if pv == cv:
                    continue

                cell = ws_current.cell(row=crow[0].row, column=c_idx)
                try:
                    pnum = float(str(pv).replace('%', ''))
                    cnum = float(str(cv).replace('%', ''))

                    if cnum > pnum:
                        cell.fill = green_fill
                        cell.value = f"{pnum:.2f}% → {cnum:.2f}% (Increased)"
                    else:
                        cell.fill = red_fill
                        cell.value = f"{pnum:.2f}% → {cnum:.2f}% (Decreased)"
                except Exception:
                    logger.error(
                        "[MRUM] Non-numeric '%s': %s vs %s", col, pv, cv
                    )

        # New entries
        for key, crow in current_map.items():
            if key not in previous_map:
                r = ws_current.max_row + 1
                for i, c in enumerate(crow, 1):
                    nc = ws_current.cell(row=r, column=i, value=c.value)
                    nc.fill = added_fill

    except Exception as e:
        logger.error(
            "[MRUM] Error in compare_overallassessmentmrum: %s", e, exc_info=True
        )
        raise


# ==============================
# MRUM SHEET DISPATCHER
# ==============================

SHEET_COMPARERS_MRUM = {
    "Analysis": compare_analysis_mrum,
    "NetworkRequestsMRUM": compare_networkrequestsmrum,
    "HealthRulesAndAlertingMRUM": compare_healthrulesandalertingmrum,
    "OverallAssessmentMRUM": compare_overallassessmentmrum,
}


def compare_files_other_sheets_mrum(
    previous_file_path: str,
    current_file_path: str,
    output_file_path: str,
) -> None:
    """
    MRUM sheet dispatcher using a sheet → function mapping.
    Mirrors the APM/BRUM dispatcher.
    """
    try:
        logger.info("[MRUM] compare_files_other_sheets_mrum starting: %s, %s -> %s",
                    previous_file_path, current_file_path, output_file_path)

        wb_previous = load_workbook(previous_file_path)
        wb_current = load_workbook(current_file_path)

        for sheet_name in wb_current.sheetnames:
            if sheet_name == "Summary":
                # Summary is handled separately by copy_summary_to_result
                continue

            if sheet_name not in wb_previous.sheetnames:
                logger.warning("[MRUM] Sheet '%s' missing in previous workbook.", sheet_name)
                continue

            ws_prev = wb_previous[sheet_name]
            ws_curr = wb_current[sheet_name]

            logger.debug("[MRUM] Processing sheet: %s", sheet_name)

            compare_func = SHEET_COMPARERS_MRUM.get(sheet_name)
            if compare_func:
                compare_func(ws_prev, ws_curr)
            else:
                logger.warning("[MRUM] No comparer defined for sheet: %s", sheet_name)

        wb_current.save(output_file_path)
        logger.info("[MRUM] Comparison results saved to: %s", output_file_path)

    except Exception as e:
        logger.error("[MRUM] Error in compare_files_other_sheets_mrum: %s", e, exc_info=True)
        raise
